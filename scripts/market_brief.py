"""Send market-brief email via Resend. Called from GitHub Actions cron."""
from __future__ import annotations

import argparse
import base64
import datetime as dt
import io
import math
import os
import sys
import urllib.request
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import requests
import yaml
import yfinance as yf


BUCKETS = {
    "Indices": ["^GSPC", "^NDX", "^STOXX50E", "^FTSE", "^N225", "^HSI"],
    "FX": ["EURUSD=X", "GBPUSD=X", "USDJPY=X", "EURGBP=X", "DX-Y.NYB"],
    "Commodities": ["CL=F", "GC=F", "HG=F", "NG=F"],
    "Single Stocks": ["NVDA", "AAPL", "MSFT", "TSLA", "ASML.AS", "MC.PA"],
}

# SX5E constituents — mirror of src/config/universe.yaml (kept here because
# the workflow only checks out the strats repo). Update when the index rebalances.
SX5E_TICKERS = [
    # DE
    "ADS.DE", "AIR.DE", "ALV.DE", "BAS.DE", "BAYN.DE", "BMW.DE", "DBK.DE", "DB1.DE",
    "DTE.DE", "DHL.DE", "EOAN.DE", "HEN3.DE", "IFX.DE", "MBG.DE", "MRK.DE", "MUV2.DE",
    "SAP.DE", "SIE.DE", "VOW3.DE",
    # FR
    "AI.PA", "CS.PA", "BNP.PA", "EN.PA", "CAP.PA", "SU.PA", "GLE.PA", "BN.PA", "EL.PA",
    "KER.PA", "OR.PA", "MC.PA", "ORA.PA", "RI.PA", "SAN.PA", "SAF.PA", "SGO.PA", "DG.PA",
    "VIV.PA", "TTE.PA",
    # NL / IT / ES / IE / FI / BE
    "ASML.AS", "INGA.AS", "PHIA.AS",
    "ENEL.MI", "ENI.MI", "ISP.MI", "UCG.MI",
    "SAN.MC", "BBVA.MC", "IBE.MC", "ITX.MC",
    "CRH.L", "NOKIA.HE", "ABI.BR",
]

# (display label, source, source-specific code)
YIELDS: list[tuple[str, str, str]] = [
    ("US 3M",  "yahoo",      "^IRX"),
    ("US 5Y",  "yahoo",      "^FVX"),
    ("US 10Y", "yahoo",      "^TNX"),
    ("US 30Y", "yahoo",      "^TYX"),
    ("UK 10Y", "boe",        "IUDMNZC"),
    ("DE 10Y", "bundesbank", "BBSIS/D.I.ZST.ZI.EUR.S1311.B.A604.R10XX.R.A.A._Z._Z.A"),
]

MAG7 = ["AAPL", "MSFT", "GOOGL", "AMZN", "NVDA", "META", "TSLA"]
SLOT_LABELS = {"am": "AM", "midday": "Midday", "pm": "PM"}

WATCHLIST_PATH = Path(__file__).parent / "watchlist.yaml"

# Yahoo option chains are US-listed only; non-US names are silently skipped.
IV_SECTORS: dict[str, list[str]] = {
    "Mega-Tech":   ["AAPL", "MSFT", "NVDA", "META", "GOOGL"],
    "Semis":       ["AVGO", "AMD", "TSM", "MU"],
    "Financials":  ["JPM", "GS", "BAC"],
    "Consumer":    ["AMZN", "WMT", "HD"],
    "Healthcare":  ["LLY", "UNH"],
    "Energy":      ["XOM", "CVX"],
    "ETFs":        ["SPY", "QQQ", "IWM"],
}
IV_TENORS_DAYS = [7, 30, 90, 180]
IV_TENOR_LABELS = {7: "1W", 30: "1M", 90: "3M", 180: "6M"}
IV_MIN_DAYS_TO_EXPIRY = 2

FONT = "Consolas, Menlo, 'Courier New', monospace"
AMBER = "#FA8C00"
BG = "#000000"
FG = "#E8E8E8"
DIM = "#888888"
POS = "#4CFF4C"
NEG = "#FF4C4C"
GRID = "#333333"
LINE_COLORS = [AMBER, "#00D9FF", POS, "#BE85FF", "#FFE14C", "#FF9EC6", NEG]


def fetch_closes(tickers: list[str], period: str = "15d") -> pd.DataFrame:
    data = yf.download(
        tickers, period=period, interval="1d", progress=False, auto_adjust=False,
        group_by="ticker", threads=True,
    )
    if isinstance(data.columns, pd.MultiIndex):
        closes = pd.DataFrame({t: data[t]["Close"] for t in tickers if t in data.columns.levels[0]})
    else:
        closes = data[["Close"]].rename(columns={"Close": tickers[0]})
    return closes.dropna(how="all")


def fetch_boe(code: str) -> pd.Series:
    today = dt.date.today()
    df_from = (today - dt.timedelta(days=30)).strftime("%d/%b/%Y")
    df_to = today.strftime("%d/%b/%Y")
    url = (
        f"https://www.bankofengland.co.uk/boeapps/iadb/fromshowcolumns.asp"
        f"?csv.x=yes&Datefrom={df_from}&Dateto={df_to}&SeriesCodes={code}"
        f"&CSVF=TN&UsingCodes=Y&VPD=Y&VFD=N"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    txt = urllib.request.urlopen(req, timeout=15).read().decode()
    rows: list[tuple[pd.Timestamp, float]] = []
    for line in txt.splitlines()[1:]:
        parts = line.split(",")
        if len(parts) < 2:
            continue
        try:
            d = pd.Timestamp(dt.datetime.strptime(parts[0].strip(), "%d %b %Y"))
            rows.append((d, float(parts[1])))
        except ValueError:
            continue
    return pd.Series({d: v for d, v in rows}).sort_index() if rows else pd.Series(dtype=float)


def fetch_bundesbank(key: str) -> pd.Series:
    url = f"https://api.statistiken.bundesbank.de/rest/data/{key}?format=csv&lastNObservations=20"
    txt = urllib.request.urlopen(url, timeout=15).read().decode()
    rows: list[tuple[pd.Timestamp, float]] = []
    for line in txt.splitlines()[2:]:
        parts = line.split(";")
        if len(parts) < 2:
            continue
        try:
            d = pd.Timestamp(parts[0].strip())
            v = float(parts[1].replace(",", "."))
            rows.append((d, v))
        except ValueError:
            continue
    return pd.Series({d: v for d, v in rows}).sort_index() if rows else pd.Series(dtype=float)


def fetch_yields() -> pd.DataFrame:
    out: dict[str, pd.Series] = {}
    yahoo_codes = [c for _, s, c in YIELDS if s == "yahoo"]
    if yahoo_codes:
        closes = fetch_closes(yahoo_codes)
        for label, src, code in YIELDS:
            if src == "yahoo" and code in closes.columns:
                out[label] = closes[code]
    for label, src, code in YIELDS:
        if src == "boe":
            try:
                s = fetch_boe(code)
                if not s.empty:
                    out[label] = s
            except Exception as e:
                print(f"BoE fetch failed for {code}: {e}", file=sys.stderr)
        elif src == "bundesbank":
            try:
                s = fetch_bundesbank(code)
                if not s.empty:
                    out[label] = s
            except Exception as e:
                print(f"Bundesbank fetch failed for {code}: {e}", file=sys.stderr)
    return pd.DataFrame(out).sort_index()


def compute_row(series: pd.Series, include_wtd: bool, mode: str = "pct") -> dict:
    s = series.dropna()
    if len(s) < 2:
        return {"last": None, "daily": None, "wtd": None}
    last = float(s.iloc[-1])
    prior = float(s.iloc[-2])
    daily = (last - prior) * 100 if mode == "bps" else (last / prior - 1) * 100
    wtd = None
    if include_wtd:
        last_date = s.index[-1].date()
        days_since_mon = last_date.weekday()
        prior_fri = last_date - dt.timedelta(days=days_since_mon + 3)
        mask = s.index.date <= prior_fri
        base = float(s[mask].iloc[-1]) if mask.any() else float(s.iloc[0])
        wtd = (last - base) * 100 if mode == "bps" else (last / base - 1) * 100
    return {"last": last, "daily": daily, "wtd": wtd}


def fmt_pct(v: float | None) -> str:
    if v is None:
        return f'<span style="color:{DIM}">n/a</span>'
    color = NEG if v < 0 else POS
    weight = "bold" if abs(v) > 1 else "normal"
    return f'<span style="color:{color};font-weight:{weight}">{v:+.2f}%</span>'


def fmt_bps(v: float | None) -> str:
    if v is None:
        return f'<span style="color:{DIM}">n/a</span>'
    color = NEG if v < 0 else POS
    weight = "bold" if abs(v) > 10 else "normal"
    return f'<span style="color:{color};font-weight:{weight}">{v:+.1f} bps</span>'


def fmt_last(v: float | None, mode: str = "pct") -> str:
    if v is None:
        return f'<span style="color:{DIM}">n/a</span>'
    if mode == "bps":
        return f"{v:.2f}%"
    return f"{v:,.2f}"


def build_table(title: str, rows: list[tuple[str, dict]], include_wtd: bool, mode: str = "pct") -> str:
    change_head = "&#916; BPS" if mode == "bps" else "DAILY %"
    wtd_head = "WTD &#916;" if mode == "bps" else "WTD %"
    cols = ["TICKER", "LAST", change_head] + ([wtd_head] if include_wtd else [])
    th_style = (
        f"text-align:right;padding:4px 10px;color:{AMBER};"
        f"border-bottom:1px solid {AMBER};font-weight:normal;letter-spacing:0.5px"
    )
    th_left = th_style.replace("text-align:right", "text-align:left")
    head = "".join(
        f"<th style=\"{th_left if i == 0 else th_style}\">{c}</th>"
        for i, c in enumerate(cols)
    )
    fmt_change = fmt_bps if mode == "bps" else fmt_pct
    body_rows = []
    for ticker, r in rows:
        td_num = f"padding:2px 10px;text-align:right;border-bottom:1px solid {GRID}"
        td_txt = f"padding:2px 10px;text-align:left;border-bottom:1px solid {GRID};color:{FG}"
        tds = [
            f'<td style="{td_txt}">{ticker}</td>',
            f'<td style="{td_num};color:{FG}">{fmt_last(r["last"], mode)}</td>',
            f'<td style="{td_num}">{fmt_change(r["daily"])}</td>',
        ]
        if include_wtd:
            tds.append(f'<td style="{td_num}">{fmt_change(r["wtd"])}</td>')
        body_rows.append(f"<tr>{''.join(tds)}</tr>")
    return (
        f'<div style="color:{AMBER};font-family:{FONT};font-size:13px;'
        f'margin:20px 0 4px;letter-spacing:1px">&#9632; {title.upper()}</div>'
        f'<table cellspacing="0" cellpadding="0" style="font-family:{FONT};'
        f'font-size:13px;border-collapse:collapse;color:{FG};width:100%">'
        f"<thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"
    )


def build_chart(series_by_name: dict[str, pd.Series], mode: str = "pct", days: int = 6) -> str:
    """Return inline <img> tag with dark/amber multi-line chart of last `days` observations."""
    fig, ax = plt.subplots(figsize=(6.4, 2.1), facecolor=BG)
    ax.set_facecolor(BG)
    for spine in ax.spines.values():
        spine.set_color(AMBER)
        spine.set_linewidth(0.5)
    ax.tick_params(colors=FG, labelsize=7, length=0)
    ax.grid(color=GRID, linestyle="-", linewidth=0.3, alpha=0.6)
    ax.axhline(0, color=FG, linewidth=0.3, alpha=0.5)

    drawn = 0
    max_len = 0
    for i, (name, s) in enumerate(series_by_name.items()):
        s = s.dropna().tail(days)
        if len(s) < 2:
            continue
        y = (s - s.iloc[0]) * 100 if mode == "bps" else (s / s.iloc[0] - 1) * 100
        color = LINE_COLORS[drawn % len(LINE_COLORS)]
        x = list(range(len(y)))
        ax.plot(x, y.values, color=color, linewidth=1.2)
        ax.text(x[-1] + 0.15, float(y.iloc[-1]), name,
                color=color, fontsize=7, va="center", family="monospace")
        drawn += 1
        max_len = max(max_len, len(y))

    if drawn == 0:
        plt.close(fig)
        return ""

    ax.set_xticks([])
    ax.set_xlim(-0.3, max_len - 1 + 1.8)
    ax.set_ylabel("bps" if mode == "bps" else "%", color=FG, fontsize=7)

    fig.tight_layout(pad=0.2)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", facecolor=BG, bbox_inches="tight", dpi=70)
    plt.close(fig)

    # Re-encode as indexed (8-bit) PNG — the chart uses a tiny palette so this
    # cuts size ~4-5x with no visible quality loss. Keeps us under Gmail's
    # ~102KB clip threshold.
    try:
        from PIL import Image
        im = Image.open(buf).convert("RGB").quantize(colors=32, method=Image.Quantize.MEDIANCUT)
        out = io.BytesIO()
        im.save(out, format="PNG", optimize=True)
        png_bytes = out.getvalue()
    except Exception:
        png_bytes = buf.getvalue()

    b64 = base64.b64encode(png_bytes).decode()
    return (
        f'<img src="data:image/png;base64,{b64}" alt="" '
        f'style="width:100%;max-width:640px;display:block;margin:6px 0 0">'
    )


def biggest_mover(all_rows: list[tuple[str, dict]]) -> tuple[str, float] | None:
    valid = [(t, r["daily"]) for t, r in all_rows if r["daily"] is not None]
    if not valid:
        return None
    return max(valid, key=lambda x: abs(x[1]))


def build_sx5e_leaderboard(price_rows: dict[str, dict], include_wtd: bool, top_n: int = 5) -> str:
    """Top-N / bottom-N daily movers within the SX5E universe."""
    ranked = [
        (t, price_rows[t]) for t in SX5E_TICKERS
        if t in price_rows and price_rows[t]["daily"] is not None
    ]
    if not ranked:
        return ""
    ranked.sort(key=lambda x: x[1]["daily"], reverse=True)
    leaders = ranked[:top_n]
    laggards = list(reversed(ranked[-top_n:]))

    header = (
        f'<div style="color:{AMBER};font-family:{FONT};font-size:13px;'
        f'margin:20px 0 4px;letter-spacing:1px">&#9632; SX5E MOVERS</div>'
    )
    return header + build_table("Leaders", leaders, include_wtd, "pct") + build_table("Laggards", laggards, include_wtd, "pct")


def load_watchlist() -> list[str]:
    if not WATCHLIST_PATH.exists():
        return []
    with WATCHLIST_PATH.open() as f:
        data = yaml.safe_load(f) or {}
    return [str(t).strip() for t in (data.get("tickers") or []) if str(t).strip()]


def fetch_iv_data(ticker: str) -> dict:
    """Pull option chains once; return spot, ATM term-structure curve, and per-expiry
    smile data (calls/puts strike+iv) so both term-structure and delta-based skew
    can be computed without duplicate HTTP calls."""
    tk = yf.Ticker(ticker)
    try:
        fi = tk.fast_info
        spot = float(fi.get("last_price") or fi.get("lastPrice"))
    except Exception:
        try:
            spot = float(tk.history(period="1d")["Close"].iloc[-1])
        except Exception:
            return {"spot": None, "atm": pd.DataFrame(columns=["days", "iv"]), "chains": {}}
    today = dt.date.today()
    atm_rows: list[tuple[int, float]] = []
    chains: dict[int, dict] = {}
    for exp in tk.options:
        try:
            days = (dt.datetime.strptime(exp, "%Y-%m-%d").date() - today).days
        except ValueError:
            continue
        if days < IV_MIN_DAYS_TO_EXPIRY:
            continue
        try:
            chain = tk.option_chain(exp)
        except Exception:
            continue
        calls = chain.calls[["strike", "impliedVolatility"]].dropna()
        puts = chain.puts[["strike", "impliedVolatility"]].dropna()
        if calls.empty or puts.empty:
            continue
        k_call = calls.iloc[(calls["strike"] - spot).abs().argsort()[:1]]
        k_put = puts.iloc[(puts["strike"] - spot).abs().argsort()[:1]]
        ivs = [
            float(k_call["impliedVolatility"].iloc[0]),
            float(k_put["impliedVolatility"].iloc[0]),
        ]
        ivs = [x for x in ivs if 0.02 < x < 3.0]
        if ivs:
            atm_rows.append((days, float(np.mean(ivs))))
        chains[days] = {"calls": calls, "puts": puts}
    atm = pd.DataFrame(atm_rows, columns=["days", "iv"]).sort_values("days").reset_index(drop=True)
    return {"spot": spot, "atm": atm, "chains": chains}


def _iv_at_delta(side: pd.DataFrame, is_call: bool, spot: float, T: float, target: float) -> float:
    """Interpolate IV vs Black-Scholes delta to hit `target`. r hardcoded to 5%
    (matters negligibly for delta)."""
    if T <= 0 or side.empty or spot is None or spot <= 0:
        return float("nan")
    deltas: list[float] = []
    ivs: list[float] = []
    for _, row in side.iterrows():
        iv = float(row["impliedVolatility"])
        K = float(row["strike"])
        if not (0.02 < iv < 3.0) or K <= 0:
            continue
        try:
            d1 = (math.log(spot / K) + (0.05 + 0.5 * iv * iv) * T) / (iv * math.sqrt(T))
        except (ValueError, ZeroDivisionError):
            continue
        d = 0.5 * (1 + math.erf(d1 / math.sqrt(2)))
        if not is_call:
            d -= 1
        deltas.append(d)
        ivs.append(iv)
    if len(deltas) < 2:
        return float("nan")
    pairs = sorted(zip(deltas, ivs), key=lambda p: p[0])
    ds, vs = [p[0] for p in pairs], [p[1] for p in pairs]
    if target <= ds[0]:
        return vs[0]
    if target >= ds[-1]:
        return vs[-1]
    return float(np.interp(target, ds, vs))


def compute_rv20(closes: pd.Series) -> float:
    """Annualised 20-day close-to-close realised vol (decimal)."""
    s = closes.dropna()
    if len(s) < 21:
        return float("nan")
    log_ret = np.log(s / s.shift(1)).dropna()
    if len(log_ret) < 20:
        return float("nan")
    return float(log_ret.tail(20).std() * math.sqrt(252))


def compute_return_moments(closes: pd.Series, window: int = 60) -> dict:
    """Skewness and excess kurtosis of the last `window` daily log returns.
    pandas .kurt() is Fisher (excess), so normal = 0."""
    s = closes.dropna()
    if len(s) < window + 1:
        return {"skew": float("nan"), "kurt": float("nan")}
    log_ret = np.log(s / s.shift(1)).dropna().tail(window)
    if len(log_ret) < window:
        return {"skew": float("nan"), "kurt": float("nan")}
    return {"skew": float(log_ret.skew()), "kurt": float(log_ret.kurt())}


def compute_pct_ranks(closes: pd.Series, iv_1m: float) -> dict:
    """
    Percentile ranks (0–100) using trailing 1-year realized distributions as proxies:
    iv_pct:  current 1M IV vs rolling-20d-RV history.
    rr_pct:  current 60d realized skew vs rolling-60d-skew history.
    bf_pct:  current 60d realized excess kurtosis vs rolling-60d-kurt history.
    """
    s = closes.dropna()
    result = {"iv_pct": float("nan"), "rr_pct": float("nan"), "bf_pct": float("nan")}
    if len(s) < 22:
        return result
    log_ret = np.log(s / s.shift(1)).dropna()

    if not math.isnan(iv_1m) and len(log_ret) >= 20:
        rv_hist = log_ret.rolling(20).std().dropna().values * math.sqrt(252)
        if len(rv_hist) >= 10:
            result["iv_pct"] = float(100 * np.mean(rv_hist <= iv_1m))

    if len(log_ret) >= 60:
        skew_hist = log_ret.rolling(60).skew().dropna()
        if len(skew_hist) >= 10:
            current_skew = float(log_ret.tail(60).skew())
            if not math.isnan(current_skew):
                result["rr_pct"] = float(100 * np.mean(skew_hist.values <= current_skew))

        kurt_hist = log_ret.rolling(60).kurt().dropna()
        if len(kurt_hist) >= 10:
            current_kurt = float(log_ret.tail(60).kurt())
            if not math.isnan(current_kurt):
                result["bf_pct"] = float(100 * np.mean(kurt_hist.values <= current_kurt))

    return result


def compute_skew_1m(data: dict) -> dict:
    """Returns {rr, bf} (25-delta risk reversal, butterfly) at ~1M using the
    nearest available expiry in [15, 60] days. Values in vol points (not pct)."""
    chains = data.get("chains") or {}
    if not chains:
        return {"rr": float("nan"), "bf": float("nan")}
    candidates = [d for d in chains if 15 <= d <= 60]
    if not candidates:
        return {"rr": float("nan"), "bf": float("nan")}
    best = min(candidates, key=lambda d: abs(d - 30))
    T = best / 365.0
    spot = data["spot"]
    c_iv = _iv_at_delta(chains[best]["calls"], True, spot, T, 0.25)
    p_iv = _iv_at_delta(chains[best]["puts"], False, spot, T, -0.25)
    atm_1m = interp_iv(data["atm"], 30)
    if math.isnan(c_iv) or math.isnan(p_iv):
        return {"rr": float("nan"), "bf": float("nan")}
    rr = p_iv - c_iv
    bf = 0.5 * (c_iv + p_iv) - atm_1m if not math.isnan(atm_1m) else float("nan")
    return {"rr": rr, "bf": bf}


def interp_iv(curve: pd.DataFrame, target_days: int) -> float:
    if curve.empty:
        return float("nan")
    d = curve["days"].to_numpy()
    iv = curve["iv"].to_numpy()
    if target_days <= d[0]:
        return float(iv[0])
    if target_days >= d[-1]:
        return float(iv[-1])
    tv = (iv ** 2) * d
    tv_t = np.interp(target_days, d, tv)
    return float(np.sqrt(tv_t / target_days))


def _fmt_pct_rank(v, td_num: str, td_na: str, high_is_bad: bool = False) -> str:
    """Render a 0–100 percentile rank. Extremes (<20 / >80) are coloured."""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return td_na
    vi = round(v)
    if vi > 80:
        color = NEG if high_is_bad else AMBER
    elif vi < 20:
        color = POS if high_is_bad else AMBER
    else:
        color = FG
    return f'<td style="{td_num};color:{color}">{vi}%</td>'


def _build_iv_table(surface: pd.DataFrame) -> str:
    """surface columns: sector, ticker, 1W, 1M, 3M, 6M, rr, bf, rv20, ivrv, move1d, rskew, rkurt, iv_pct, rr_pct, bf_pct.
    IV/RV values stored as decimals; displayed as vol points (×100)."""
    tenor_cols = [IV_TENOR_LABELS[d] for d in IV_TENORS_DAYS]
    cols = ["TICKER"] + tenor_cols + ["6M-1W", "IV%", "1M RR", "RR%", "1M BF", "BF%", "20d RV", "IV-RV", "1d Mv", "Skew60", "Kurt60"]
    th_style = (
        f"text-align:right;padding:4px 10px;color:{AMBER};"
        f"border-bottom:1px solid {AMBER};font-weight:normal;letter-spacing:0.5px"
    )
    th_left = th_style.replace("text-align:right", "text-align:left")
    head = "".join(
        f'<th style="{th_left if i == 0 else th_style}">{c}</th>'
        for i, c in enumerate(cols)
    )
    body_rows: list[str] = []
    ncols = len(cols)
    for sector, grp in surface.groupby("sector", sort=False):
        body_rows.append(
            f'<tr><td colspan="{ncols}" style="padding:6px 10px 2px;'
            f'color:{AMBER};font-size:11px;letter-spacing:1px;'
            f'border-bottom:1px solid {GRID}">{sector.upper()}</td></tr>'
        )
        for _, r in grp.iterrows():
            td_num = f"padding:2px 10px;text-align:right;border-bottom:1px solid {GRID};color:{FG}"
            td_txt = f"padding:2px 10px;text-align:left;border-bottom:1px solid {GRID};color:{FG}"
            td_na = f'<td style="{td_num};color:{DIM}">n/a</td>'
            tds = [f'<td style="{td_txt}">{r["ticker"]}</td>']
            for c in tenor_cols:
                v = r[c]
                tds.append(f'<td style="{td_num}">{v * 100:.1f}</td>' if pd.notna(v) else td_na)
            slope = (r["6M"] - r["1W"]) * 100 if pd.notna(r["6M"]) and pd.notna(r["1W"]) else None
            if slope is None:
                tds.append(td_na)
            else:
                tds.append(f'<td style="{td_num};color:{NEG if slope < 0 else POS}">{slope:+.1f}</td>')
            # IV% — high pct = expensive vol (red)
            tds.append(_fmt_pct_rank(r.get("iv_pct", float("nan")), td_num, td_na, high_is_bad=True))
            rr = r.get("rr")
            if pd.isna(rr):
                tds.append(td_na)
            else:
                rr_pts = rr * 100
                tds.append(f'<td style="{td_num};color:{NEG if rr_pts < 0 else POS}">{rr_pts:+.1f}</td>')
            # RR% — extremes amber (unusual skew regime)
            tds.append(_fmt_pct_rank(r.get("rr_pct", float("nan")), td_num, td_na, high_is_bad=False))
            bf = r.get("bf")
            tds.append(f'<td style="{td_num}">{bf * 100:+.1f}</td>' if pd.notna(bf) else td_na)
            # BF% — extremes amber (unusual kurtosis regime)
            tds.append(_fmt_pct_rank(r.get("bf_pct", float("nan")), td_num, td_na, high_is_bad=False))
            rv = r.get("rv20")
            tds.append(f'<td style="{td_num}">{rv * 100:.1f}</td>' if pd.notna(rv) else td_na)
            ivrv = r.get("ivrv")
            if pd.isna(ivrv):
                tds.append(td_na)
            else:
                ivrv_pts = ivrv * 100
                tds.append(f'<td style="{td_num};color:{NEG if ivrv_pts < 0 else POS}">{ivrv_pts:+.1f}</td>')
            mv = r.get("move1d")
            tds.append(f'<td style="{td_num}">{mv * 100:.2f}%</td>' if pd.notna(mv) else td_na)
            rskew = r.get("rskew")
            if pd.isna(rskew):
                tds.append(td_na)
            else:
                tds.append(f'<td style="{td_num};color:{NEG if rskew < 0 else POS}">{rskew:+.2f}</td>')
            rkurt = r.get("rkurt")
            tds.append(f'<td style="{td_num}">{rkurt:+.1f}</td>' if pd.notna(rkurt) else td_na)
            body_rows.append(f"<tr>{''.join(tds)}</tr>")
    return (
        f'<table cellspacing="0" cellpadding="0" style="font-family:{FONT};'
        f'font-size:13px;border-collapse:collapse;color:{FG};width:100%;margin-top:6px">'
        f"<thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"
    )


def build_iv_section(watchlist_tickers: list[str]) -> tuple[str, pd.DataFrame]:
    sectors: dict[str, list[str]] = {k: list(v) for k, v in IV_SECTORS.items()}
    us_watchlist = [t for t in watchlist_tickers if "." not in t and "-" not in t]
    if us_watchlist:
        sectors["Watchlist"] = us_watchlist

    all_tickers = sorted({t for ts in sectors.values() for t in ts})
    data_by_ticker: dict[str, dict] = {}
    for t in all_tickers:
        try:
            d = fetch_iv_data(t)
            if not d["atm"].empty:
                data_by_ticker[t] = d
        except Exception as e:
            print(f"IV fetch failed for {t}: {e}", file=sys.stderr)

    if not data_by_ticker:
        return "", pd.DataFrame()

    rv_closes = fetch_closes(list(data_by_ticker.keys()), period="1y")

    tenor_cols = [IV_TENOR_LABELS[d] for d in IV_TENORS_DAYS]
    rows: list[dict] = []
    for sector, tickers in sectors.items():
        for t in tickers:
            if t not in data_by_ticker:
                continue
            d = data_by_ticker[t]
            skew = compute_skew_1m(d)
            atm_1m = interp_iv(d["atm"], 30)
            series = rv_closes[t] if t in rv_closes.columns else pd.Series(dtype=float)
            rv20 = compute_rv20(series)
            moments = compute_return_moments(series, window=60)
            pcts = compute_pct_ranks(series, atm_1m)
            ivrv = atm_1m - rv20 if not (math.isnan(atm_1m) or math.isnan(rv20)) else float("nan")
            move1d = atm_1m * math.sqrt(1 / 252) if not math.isnan(atm_1m) else float("nan")
            rows.append({
                "sector": sector, "ticker": t,
                **{IV_TENOR_LABELS[tn]: interp_iv(d["atm"], tn) for tn in IV_TENORS_DAYS},
                "rr": skew["rr"], "bf": skew["bf"],
                "rv20": rv20, "ivrv": ivrv, "move1d": move1d,
                "rskew": moments["skew"], "rkurt": moments["kurt"],
                "iv_pct": pcts["iv_pct"], "rr_pct": pcts["rr_pct"], "bf_pct": pcts["bf_pct"],
            })
    surface = pd.DataFrame(
        rows,
        columns=["sector", "ticker"] + tenor_cols + ["rr", "bf", "rv20", "ivrv", "move1d", "rskew", "rkurt", "iv_pct", "rr_pct", "bf_pct"],
    )
    if surface.empty:
        return "", pd.DataFrame()

    header = (
        f'<div style="color:{AMBER};font-family:{FONT};font-size:13px;'
        f'margin:20px 0 4px;letter-spacing:1px">&#9632; VOL SURFACE</div>'
    )
    return header + _build_iv_table(surface), surface


def build_email(slot: str) -> tuple[str, str, dict]:
    """Fetch data and build the email. Returns (subject, html, excel_data). No network send."""
    include_wtd = slot == "pm"
    label = SLOT_LABELS[slot]

    watchlist_tickers = load_watchlist()

    price_tickers = sorted(
        {t for b in BUCKETS.values() for t in b} | set(MAG7) | set(watchlist_tickers) | set(SX5E_TICKERS)
    )
    closes = fetch_closes(price_tickers)
    yields_df = fetch_yields()

    price_rows: dict[str, dict] = {}
    failures: list[str] = []
    for t in price_tickers:
        if t not in closes.columns:
            failures.append(t)
            price_rows[t] = {"last": None, "daily": None, "wtd": None}
            continue
        row = compute_row(closes[t], include_wtd, "pct")
        if row["last"] is None:
            failures.append(t)
        price_rows[t] = row

    yield_rows: dict[str, dict] = {}
    for ylabel, _, _ in YIELDS:
        if ylabel in yields_df.columns:
            row = compute_row(yields_df[ylabel], include_wtd, "bps")
        else:
            row = {"last": None, "daily": None, "wtd": None}
        if row["last"] is None:
            failures.append(ylabel)
        yield_rows[ylabel] = row

    sections: list[str] = []
    price_flat: list[tuple[str, dict]] = []

    # Indices
    idx_tickers = BUCKETS["Indices"]
    rows = [(t, price_rows[t]) for t in idx_tickers]
    price_flat.extend(rows)
    sections.append(
        build_table("Indices", rows, include_wtd, "pct")
        + build_chart({t: closes[t] for t in idx_tickers if t in closes.columns}, "pct")
    )

    # Yields
    y_rows = [(lbl, yield_rows[lbl]) for lbl, _, _ in YIELDS]
    sections.append(
        build_table("Yields", y_rows, include_wtd, "bps")
        + build_chart({lbl: yields_df[lbl] for lbl, _, _ in YIELDS if lbl in yields_df.columns}, "bps")
    )

    # FX
    fx_tickers = BUCKETS["FX"]
    rows = [(t, price_rows[t]) for t in fx_tickers]
    price_flat.extend(rows)
    sections.append(
        build_table("FX", rows, include_wtd, "pct")
        + build_chart({t: closes[t] for t in fx_tickers if t in closes.columns}, "pct")
    )

    # Commodities
    comm_tickers = BUCKETS["Commodities"]
    rows = [(t, price_rows[t]) for t in comm_tickers]
    price_flat.extend(rows)
    sections.append(
        build_table("Commodities", rows, include_wtd, "pct")
        + build_chart({t: closes[t] for t in comm_tickers if t in closes.columns}, "pct")
    )

    # Single Stocks
    ss_tickers = BUCKETS["Single Stocks"]
    rows = [(t, price_rows[t]) for t in ss_tickers]
    price_flat.extend(rows)
    sections.append(
        build_table("Single Stocks", rows, include_wtd, "pct")
        + build_chart({t: closes[t] for t in ss_tickers if t in closes.columns}, "pct")
    )

    # SX5E movers
    sx5e_html = build_sx5e_leaderboard(price_rows, include_wtd)
    if sx5e_html:
        sections.append(sx5e_html)

    # Watchlist
    if watchlist_tickers:
        wl_rows_tbl = [(t, price_rows.get(t, {"last": None, "daily": None, "wtd": None})) for t in watchlist_tickers]
        price_flat.extend(wl_rows_tbl)
        sections.append(
            build_table("Watchlist", wl_rows_tbl, include_wtd, "pct")
            + build_chart({t: closes[t] for t in watchlist_tickers if t in closes.columns}, "pct")
        )

    # Custom factors: Mag7 equal-weight
    mag7_daily = [price_rows[t]["daily"] for t in MAG7 if price_rows[t]["daily"] is not None]
    mag7_row = {
        "last": None,
        "daily": sum(mag7_daily) / len(mag7_daily) if mag7_daily else None,
        "wtd": None,
    }
    if include_wtd:
        mag7_wtd = [price_rows[t]["wtd"] for t in MAG7 if price_rows[t]["wtd"] is not None]
        mag7_row["wtd"] = sum(mag7_wtd) / len(mag7_wtd) if mag7_wtd else None
    sections.append(build_table("Custom Factors", [("Mag7 (eq-wt)", mag7_row)], include_wtd, "pct"))

    # IV term structure — shown on all slots.
    iv_html, surface_df = build_iv_section(watchlist_tickers)
    if iv_html:
        sections.append(iv_html)

    mover = biggest_mover(price_flat)
    if mover:
        mover_color = NEG if mover[1] < 0 else POS
        commentary = (
            f'BIGGEST MOVER: <span style="color:{AMBER};font-weight:bold">{mover[0]}</span> '
            f'<span style="color:{mover_color};font-weight:bold">{mover[1]:+.2f}%</span>'
        )
    else:
        commentary = "NO VALID PRICES RETRIEVED"

    today = dt.date.today().isoformat()
    subject = f"Market brief — {label} — {today}"

    header_bar = (
        f'<div style="background:{BG};border-bottom:2px solid {AMBER};'
        f'padding:10px 14px;font-family:{FONT};font-size:13px;color:{AMBER};'
        f'letter-spacing:2px;font-weight:bold">'
        f'MARKET BRIEF &nbsp;&#9474;&nbsp; {label.upper()} &nbsp;&#9474;&nbsp; {today}'
        f'</div>'
    )
    commentary_bar = (
        f'<div style="padding:10px 14px;font-family:{FONT};font-size:13px;'
        f'color:{FG};border-bottom:1px solid {GRID}">{commentary}</div>'
    )
    footer = (
        f'<div style="padding:12px 14px;margin-top:18px;border-top:1px solid {GRID};'
        f'font-family:{FONT};font-size:11px;color:{DIM};letter-spacing:0.5px">'
        f'FAILED TICKERS: {", ".join(failures)}</div>'
        if failures else ""
    )
    html = (
        f'<table cellspacing="0" cellpadding="0" style="width:100%;background:{BG};'
        f'border-collapse:collapse"><tr><td bgcolor="{BG}" style="background:{BG};'
        f'padding:0">'
        f'<div style="background:{BG};padding:0 14px 16px">'
        f"{header_bar}{commentary_bar}"
        f'<div style="padding:0 4px">{"".join(sections)}</div>'
        f"{footer}"
        f"</div>"
        f"</td></tr></table>"
    )
    excel_data = {
        "price_rows": price_rows,
        "yield_rows": yield_rows,
        "buckets": BUCKETS,
        "sx5e_tickers": SX5E_TICKERS,
        "watchlist_tickers": watchlist_tickers,
        "mag7_row": mag7_row,
        "surface_df": surface_df,
        "include_wtd": include_wtd,
    }
    return subject, html, excel_data


def _xl_header(ws, values: list[str]) -> None:
    """Write a bold amber header row to a worksheet."""
    fill = PatternFill("solid", fgColor="FA8C00")
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=1, column=col, value=v)
        cell.font = Font(bold=True, color="000000")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def build_excel_bytes(
    price_rows: dict[str, dict],
    yield_rows: dict[str, dict],
    buckets: dict[str, list[str]],
    sx5e_tickers: list[str],
    watchlist_tickers: list[str],
    mag7_row: dict,
    surface_df: pd.DataFrame,
    include_wtd: bool,
) -> bytes:
    wb = openpyxl.Workbook()

    # ── Prices sheet ───────────────────────────────────────────────────────────
    ws_p = wb.active
    ws_p.title = "Prices"
    price_headers = ["Ticker", "Section", "Last", "Daily %"] + (["WTD %"] if include_wtd else [])
    _xl_header(ws_p, price_headers)

    def _price_rows_iter():
        for section, tickers in buckets.items():
            for t in tickers:
                r = price_rows.get(t, {})
                yield t, section, r.get("last"), r.get("daily"), r.get("wtd")
        for t in sx5e_tickers:
            r = price_rows.get(t, {})
            yield t, "SX5E", r.get("last"), r.get("daily"), r.get("wtd")
        for t in watchlist_tickers:
            r = price_rows.get(t, {})
            yield t, "Watchlist", r.get("last"), r.get("daily"), r.get("wtd")
        yield "Mag7 (eq-wt)", "Custom", None, mag7_row.get("daily"), mag7_row.get("wtd")

    for row_i, (ticker, section, last, daily, wtd) in enumerate(_price_rows_iter(), 2):
        ws_p.cell(row=row_i, column=1, value=ticker)
        ws_p.cell(row=row_i, column=2, value=section)
        ws_p.cell(row=row_i, column=3, value=round(last, 4) if last is not None else None)
        ws_p.cell(row=row_i, column=4, value=round(daily, 4) if daily is not None else None)
        if include_wtd:
            ws_p.cell(row=row_i, column=5, value=round(wtd, 4) if wtd is not None else None)
    for col in ws_p.columns:
        ws_p.column_dimensions[col[0].column_letter].width = 16

    # ── Yields sheet ──────────────────────────────────────────────────────────
    ws_y = wb.create_sheet("Yields")
    yield_headers = ["Tenor", "Last (%)", "Daily (bps)"] + (["WTD (bps)"] if include_wtd else [])
    _xl_header(ws_y, yield_headers)
    for row_i, (label, r) in enumerate(yield_rows.items(), 2):
        ws_y.cell(row=row_i, column=1, value=label)
        ws_y.cell(row=row_i, column=2, value=round(r["last"], 3) if r["last"] is not None else None)
        ws_y.cell(row=row_i, column=3, value=round(r["daily"], 2) if r["daily"] is not None else None)
        if include_wtd:
            ws_y.cell(row=row_i, column=4, value=round(r["wtd"], 2) if r["wtd"] is not None else None)
    for col in ws_y.columns:
        ws_y.column_dimensions[col[0].column_letter].width = 16

    # ── Vol Surface sheet ──────────────────────────────────────────────────────
    if not surface_df.empty:
        ws_v = wb.create_sheet("Vol Surface")
        vol_cols = list(surface_df.columns)
        _xl_header(ws_v, vol_cols)
        for row_i, (_, row) in enumerate(surface_df.iterrows(), 2):
            for col_i, col in enumerate(vol_cols, 1):
                v = row[col]
                if isinstance(v, float) and math.isnan(v):
                    v = None
                elif isinstance(v, float):
                    v = round(v, 6)
                ws_v.cell(row=row_i, column=col_i, value=v)
        for col in ws_v.columns:
            ws_v.column_dimensions[col[0].column_letter].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--slot", choices=["am", "midday", "pm"], required=True)
    args = ap.parse_args()

    subject, html, excel_data = build_email(args.slot)

    xlsx_bytes = build_excel_bytes(**excel_data)
    xlsx_b64 = base64.b64encode(xlsx_bytes).decode()
    today = dt.date.today().isoformat()
    xlsx_filename = f"market_brief_{today}_{args.slot}.xlsx"

    api_key = os.environ["RESEND_API_KEY"]
    r = requests.post(
        "https://api.resend.com/emails",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "from": "onboarding@resend.dev",
            "to": ["bailey.arm.business@gmail.com"],
            "subject": subject,
            "html": html,
            "attachments": [{"filename": xlsx_filename, "content": xlsx_b64}],
        },
        timeout=30,
    )
    if r.status_code != 200:
        print(f"Resend error {r.status_code}: {r.text}", file=sys.stderr)
        return 1
    print(f"sent: {r.json().get('id')}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
